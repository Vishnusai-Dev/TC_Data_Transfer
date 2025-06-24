
import streamlit as st
import pandas as pd
import os
import zipfile
from io import BytesIO
from openpyxl import load_workbook

st.title("Excel/CSV Data Transfer App")

uploaded_master = st.file_uploader("Upload Master Excel File", type=["xlsx"])
uploaded_files = st.file_uploader("Upload Target Files (Excel or CSV)", type=["xlsx", "csv"], accept_multiple_files=True)

if uploaded_master and uploaded_files:
    st.success("Files uploaded successfully")

    if st.button("Run Data Transfer"):
        try:
            log = []
            output_zip = BytesIO()
            zipf = zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED)

            df_master = pd.read_excel(uploaded_master, sheet_name="Master Data", header=3)  # Reads from row 4 (0-indexed = 3)

            for file in uploaded_files:
                try:
                    if file.name.endswith('.csv'):
                        df_template = pd.read_csv(file)
                        headers = df_template.columns.tolist()
                        file_path_col = df_master.columns[0]  # Assuming column A
                        matching_rows = df_master[df_master[file_path_col] == file.name]

                        for _, row in matching_rows.iterrows():
                            row_data = {}
                            for header in headers:
                                if header in row:
                                    row_data[header] = row[header]
                            df_template = pd.concat([df_template, pd.DataFrame([row_data])], ignore_index=True)

                        csv_data = df_template.to_csv(index=False)
                        zipf.writestr(file.name, csv_data)

                    elif file.name.endswith('.xlsx'):
                        wb = load_workbook(file)
                        ws = wb.worksheets[0]

                        # Get headers from row 4
                        headers = [cell.value for cell in ws[4]]

                        # Create a DataFrame with same headers
                        df_template = pd.DataFrame(columns=headers)

                        file_path_col = df_master.columns[0]  # Assuming column A
                        matching_rows = df_master[df_master[file_path_col] == file.name]

                        for _, row in matching_rows.iterrows():
                            row_data = {}
                            for header in headers:
                                if header in row:
                                    row_data[header] = row[header]
                            df_template = df_template.append(row_data, ignore_index=True)

                        # Append to file (after existing data)
                        for idx, row in df_template.iterrows():
                            ws.append(row.tolist())

                        temp_io = BytesIO()
                        wb.save(temp_io)
                        zipf.writestr(file.name, temp_io.getvalue())
                except Exception as e:
                    log.append((file.name, str(e)))

            zipf.close()
            st.download_button("Download Transferred Files (ZIP)", data=output_zip.getvalue(), file_name="output_transferred.zip")

            if log:
                st.write("Errors:")
                for entry in log:
                    st.error(f"{entry[0]} - {entry[1]}")
            else:
                st.success("All files processed successfully.")
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
